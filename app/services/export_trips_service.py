from .trip_history_service import TripHistoryService
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os


class ExportTripsService:
    def __init__(self, trip_history_service: TripHistoryService):
        self.trip_history_service = trip_history_service

    def export_trips_to_excel(self, company, date_from, date_to):
        trips_by_date_list, total_trips, total_price, total_passengers = self.trip_history_service.get_daily_trips(
            company,
            date_from,
            date_to,
        )
        rows = []

        filename = self.get_filename()

        for day in trips_by_date_list:
            date = day["date"]

            for trip in day["trips"]:

                # 1 row per passenger
                for p in trip.trip_passengers:
                    rows.append({
                        "Date": date,
                        "Taxi": trip.taxi,
                        "Time": trip.start_time,
                        "Society": trip.society,
                        "Price": trip.price,
                        "Passenger": p.passenger,
                        "Pickup": p.pickup,
                        "Dropoff": p.dropoff,
                    })

                # separator row (blank line)
                rows.append({
                    "Date": "─",
                    "Taxi": "─",
                    "Time": "─",
                    "Society": "─",
                    "Price": "─",
                    "Passenger": "─",
                    "Pickup": "─",
                    "Dropoff": "─",
                })

        df = pd.DataFrame(rows)
        df.to_excel(filename, index=False)
        # ---------------- STYLE PART ----------------
        wb = load_workbook(filename)
        ws = wb.active

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")

        trip_end_fill = PatternFill("solid", fgColor="E7E6E6")

        center = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # HEADER STYLE
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # COLUMN WIDTHS
        widths = [15, 12, 10, 15, 12, 20, 18, 18, 10]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        # STYLE ROWS
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = center
                cell.border = thin_border

                # separator row detection
                if cell.column == 4 and cell.value == "─":
                    for c in row:
                        c.fill = trip_end_fill
                        c.font = Font(bold=True)

        wb.save(filename)

        return filename

    def get_filename(self):

        base_url = os.path.dirname(os.path.abspath(__file__))
        export_dir = os.path.join(base_url, "exports")
        os.makedirs(export_dir, exist_ok=True)

        filename = f"trips_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = os.path.join(export_dir, filename)

        return file_path
